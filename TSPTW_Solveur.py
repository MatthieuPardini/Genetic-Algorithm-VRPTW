#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TSPTW - Solveur
"""

from openpyxl import load_workbook
import numpy as np
from math import sqrt
from pyscipopt import Model, quicksum,Expr

M=10000             # Grand nombre
limite_temps = 180  # Limite du temps de résolution en secondes

def main():
    print('Lecture des données')
    lecteur("Données/Inst20_3.xlsx")
    print("*** Résolution en cours par le solveur avec limite de temps ",limite_temps," secondes ***")
    LinMod()

def lecteur(fname):
    global wb
    wb = load_workbook(fname,data_only=True)
    global ws
    ws = wb.worksheets[0]
    global n
    n=0
    while  ws.cell(n+2,1).value is not None : 
        n=n+1
    global coX; coX = np.zeros(n,dtype=float)
    global coY; coY = np.zeros(n,dtype=float) 
    global a; a = np.zeros(n,dtype=float)
    global b; b = np.zeros(n,dtype=float)
    global penalite; penalite = np.zeros(n,dtype=float)
    global dem; dem = np.zeros(n,dtype=float)
    
    for i in range(n):
        coX[i]= ws.cell(i+2,2).value
        coY[i]= ws.cell(i+2,3).value
        a[i]= ws.cell(i+2,4).value
        b[i]= ws.cell(i+2,5).value
        penalite[i] = ws.cell(i+2,6).value
        dem[i] = ws.cell(i+2,7).value
    global capa; capa = ws.cell(2,8).value
    
    global dist; dist=np.zeros([n,n], dtype=float)
    for i in range(n):
        for j in range(n):
            dist[i][j]=sqrt((coX[i]-coX[j])**2+(coY[i]-coY[j])**2)

def LinMod():
    model = Model("TSPTW")
    x={}
    t={}
    r={} # retard sur le noeud i, 0 si pas de retard
    for i in range(n):
        t[i]=model.addVar(lb=0,ub=M,name="t[%s]"%i)
        r[i]=model.addVar(lb=0,name="r[%s]"%i)
        for j in range(n):
            x[i,j]=model.addVar(vtype="B",name="x[%s,%s]"%(i,j))
            
    model.setObjective(quicksum(x[i,j]*dist[i,j] for i in range(n) for j in range(n))+ quicksum(penalite[i]*r[i] for i in range(n)),"minimize")
    

    mycst = Expr()
    for j in range(n): 
        mycst+=x[0,j]
    model.addCons(mycst==1)
    
    for j in range(n):    
        model.addCons(quicksum(x[i,j] for i in range(n) if i!=j) ==1,"Visite %s"%j)
    for j in range(n):
        model.addCons(quicksum(x[i,j] for i in range(n)if i!=j) - quicksum(x[j,i]for i in range(n)if i!=j) ==0,"Cont %s"%j)
   
    for j in range(1,n):
        for i in range(n):
            model.addCons(t[j]>= t[i]+dist[i,j]-10000*(1-x[i,j]))
   
    model.addCons(t[0]==0)
    
    
    for j in range(n):
        model.addCons(t[j]>=a[j],"Fenêtre %s"%j)
    
    for j in range(n):
        model.addCons(r[j]>=t[j]-b[j],"Retard %s"%j)
    
    
    model.hideOutput()
    model.setParam("limits/time",limite_temps)
    
    model.optimize()

    sol = model.getBestSol()
    print("Fonction obj %s "%model.getSolObjVal(sol)) 
    print("Temps de resolution", model.getSolvingTime() )
    print("Variables :")
    for i in range(n):
        for j in range(n):
            if model.getSolVal(sol,x[i,j])>0.99:
                print(i,"->",j)
    
    
    print("Tournee commencant par depot ")
    ori=0
    dest=-1
    while dest!=0:
        dest=0
        goon=True
        while dest<n and goon :
            if(model.getSolVal(sol,x[ori,dest])>0.5) : 
                print(ori,"->",dest) 
                ori=dest
                goon=False
            else:dest+=1
   
    
    
if __name__=="__main__":
    main()