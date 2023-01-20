#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Algorithme génétique (sans contrainte de capacité)
"""
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import time
from math import sqrt
import matplotlib.pyplot as plt
from tqdm import tqdm


# Paramètres globaux
m=100                           # Taille de la population
nbgen=1600                      # Nombre de générations
prob_mut=20                     # Pourcentage : probabilité de mutation
part_heuristique = 60           # Pourcentage : part de solutions initiales créées avec heuristique

fig, ax = plt.subplots()

def main():
    lecteur('Données/Inst40_1.xlsx')
    plt.axis('off')
    ax.scatter(coX, coY)
    
    Noeuds = np.arange(n).tolist()
    
    for i, txt in enumerate(Noeuds):
        ax.annotate(txt, (coX[i], coY[i]))
        plt.scatter(coX,coY)
    
    execution_unique()
    #execution_multiple()
    for i in range(n-1):
        afficher(bsolution[nbgen-1][i],bsolution[nbgen-1][i+1])
    
def afficher(i,j):
    xx = np.linspace(min(coX[i],coX[j]),max(coX[i],coX[j]))
    if coX[i]!=coX[j]:
        a,b = coeffs_affine(coX[i], coX[j], coY[i], coY[j])
        yy = fct_affine(xx,a,b)
    else:
        yy = np.linspace(coY[i],coY[j])
    ax.plot(xx,yy,color='r')

def coeffs_affine(xa,xb,ya,yb):
    a = (yb-ya)/(xb-xa)
    b = ya - a*xa 
    return a,b

def fct_affine(x,a,b):
    return a*x+b

def lecteur(fname):
    global wb
    wb = load_workbook(fname,data_only=True)
    global ws
    ws = wb.worksheets[0]
    global n
    n=0
    while  ws.cell(n+2,1).value is not None : 
        n=n+1
    global coX; coX = np.zeros(n,dtype=float)
    global coY; coY = np.zeros(n,dtype=float) 
    global a; a = np.zeros(n,dtype=float)
    global b; b = np.zeros(n,dtype=float)
    global penalite; penalite = np.zeros(n,dtype=float)
    global dem; dem = np.zeros(n,dtype=float)
    
    for i in range(n):
        coX[i]= ws.cell(i+2,2).value
        coY[i]= ws.cell(i+2,3).value
        a[i]= ws.cell(i+2,4).value
        b[i]= ws.cell(i+2,5).value
        penalite[i] = ws.cell(i+2,6).value
        dem[i] = ws.cell(i+2,7).value
    global capa; capa = ws.cell(2,8).value
    
    print("*** Données récupérées ***")
    
    global dist; dist=np.zeros([n,n], dtype=float)
    for i in range(n):
        for j in range(n):
            dist[i][j]=sqrt((coX[i]-coX[j])**2+(coY[i]-coY[j])**2)
    print("*** Distancier créé ***")
    return

def hasard():
    route=np.arange(1,n)
    random.shuffle(route)
    route= np.insert(route,0,0) 
    return route 

def two_random_clarke_wright_tsp():
    ''' Version aléatoire de la heuristique constructive de Clarke & Wright
    Choisit aléatoirement entre les deux meilleurs regroupements de tournées'''
    #Pré-processing : calcul de tous les Sij (amélioration de regrouper la tournée finissant par i à celle commençant par j)
    s=np.zeros([n,n],dtype=float)
    for i in range(n):
        for j in range(n):
            s[i,j]=dist[0,i]+dist[j,0]-dist[i,j]
            
    # Initialisation
    tournees = []
    for i in range(n-1):
        tournees.append([0,i+1,0])
    
    for iteration in range(1,n-1):
        smax=-1; tournee_debut_best = 0; tournee_fin_best=0
        smax2=-1; tournee_debut_best2 = 0; tournee_fin_best2=0
        
        # Recherche des deux meilleures fusions :
        for t1 in range(n-iteration):
            i1 = tournees[t1][1]-1;j1=tournees[t1][-2]-1
            for t2 in range(t1+1,n-iteration):
                i2 = tournees[t2][1]-1;j2=tournees[t2][-2]-1
                #Test pour trouver le meilleur
                if s[j1,i2]>=smax:
                    smax=s[j1,i2]
                    tournee_debut_best = t1
                    tournee_fin_best = t2
                #Test pour trouver le deuxième meilleur
                if s[j1,i2]<smax and s[j1,i2]>=smax2:
                    smax2=s[j1,i2]
                    tournee_debut_best2 = t1
                    tournee_fin_best2 = t2                 
                    
                #Test pour trouver le meilleur (autre sens)
                if s[j2,i1]>=smax:
                    smax=s[j2,i1]
                    tournee_debut_best = t2
                    tournee_fin_best = t1
                #Test pour trouver le deuxième meilleur (autre sens)
                if s[j2,i1]<smax and s[j2,i1]>=smax2:
                    smax2=s[j2,i1]
                    tournee_debut_best2 = t2
                    tournee_fin_best2 = t1  
        
        # Choix aléatoire :
        a = random.randint(0,1)
        if smax2==-1 or a == 0:
            # Choix 1 : faire la meilleure fusion de tournées
            debut_choisi = tournee_debut_best
            fin_choisie = tournee_fin_best
        else:
            # Choix 2 : faire la deuxième meilleure fusion de tournées
            debut_choisi = tournee_debut_best2
            fin_choisie = tournee_fin_best2
        
        nouvelle_tournee = tournees[debut_choisi][:-1] + tournees[fin_choisie][1:]
        tournees.pop(max(debut_choisi,fin_choisie))
        tournees.pop(min(debut_choisi,fin_choisie))
        tournees.append(nouvelle_tournee)
    return tournees[0][:-1] # [:-1] pour ne pas compter le retour au dépôt dans l'expression de la solution


def initialize():
    global population, costs
    costs=np.zeros(m,dtype=float)
    population=np.zeros([m,n],dtype=int)
    #initialiser la population initiale
    
    #Partie heuristique
    for i in range(m//int(100/part_heuristique)):
        population[i]=two_random_clarke_wright_tsp()
        costs[i]=calculate_cost(population[i])
    
    #Partie hasard
    for i in range(m//int(100/part_heuristique),m):
        population[i]=hasard()
        costs[i]=calculate_cost(population[i])
   
#Calcul du coût
def calculate_cost(route):
    cost=0
    for i in range(1,len(route)):
        cost+=dist[route[i-1], route[i]]
    cost+=dist[route[-1],0]
    
    times = time_windows(route)
    
    for i in range(1,len(route)):
        if times[i] > b[route[i]]:
            cost += penalite[route[i]]*(times[i]-b[route[i]])
            
    return cost

def time_windows(route):
    times = np.zeros(n,dtype=float)
    times[0]=0
    for i in range(1,n):
        times[i] = max(times[i-1]+dist[route[i-1],route[i]],a[route[i]])
    return times



def mutation():
    for i in range(m):
        population[i]=mutation_swap(population[i])

def mutation_swap(p):
    #Choix de mutation : inversion
    tirage_aleatoire = random.randint(1,100)
    if tirage_aleatoire <= prob_mut:
        indice_gene1 = random.randint(1,n-1)
        indice_gene2 = random.randint(1,n-1)
        while indice_gene2 == indice_gene1:
            indice_gene2 = random.randint(1,n-1)
        gene1=p[indice_gene1]
        gene2=p[indice_gene2]
        p[indice_gene1]=gene2
        p[indice_gene2]=gene1
    return p

def cross_over(parent1,parent2):
    point_croisement = random.randint(1,n-2)
    enfant1 = np.zeros(n,dtype=int)
    enfant2 = np.zeros(n,dtype=int)
    visited1 = np.zeros(n,dtype=bool)
    visited2 = np.zeros(n,dtype=bool)
    
    for i in range(point_croisement):
        enfant1[i] = parent1[i]
        visited1[parent1[i]] = True
        enfant2[i] = parent2[i]
        visited2[parent2[i]]=True
    j1 = point_croisement
    j2 = point_croisement
    
    for i in range(0,n):   
        if not visited1[parent2[i]]:
            enfant1[j1] = parent2[i]
            j1 = j1+1
        if not visited2[parent1[i]]:
            enfant2[j2] = parent1[i]
            j2 = j2+1
    
    return enfant1, enfant2

def croisement_selectif():
    for z in range(0,m,2):
        family = np.zeros([4,n],dtype=int)
        family[0] = population[z] # parent1
        family[1]= population[z+1] #parent2
        family[2],family[3] = cross_over(family[0], family[1]) # enfants 1 et 2
    
        costs_family = np.zeros(4,dtype=float)
        for k in range(4):
            costs_family[k]=calculate_cost(family[k])
        #Sélection des 2 meilleurs :
        indices_meilleurs = np.argsort(costs_family)
        meilleur = family[indices_meilleurs[0]]
        deuxieme_meilleur = family[indices_meilleurs[1]]
        
        
        population[z] = meilleur
        population[z+1] = deuxieme_meilleur


    
def recherche_locale_2opt(route):
    cost_route = calculate_cost(route)
    best_cost = cost_route
    best_route = route[:]
    for i in range(1,n-2):
        for j in range(i+2,n):
            route_test = np.concatenate((route[:i],np.flip(route[i:j]),route[j:n]))
            cost_test = calculate_cost(route_test)
            if cost_test < best_cost:
                best_cost = cost_test
                best_route = route_test
    return best_route,best_cost



def shuffle_population():
    population_copie = population[:]
    indice_aleatoire = np.zeros(m,dtype=int)
    indice_utilise = np.zeros(m,dtype=bool)
    for i in range(m):
        indice_aleatoire[i]=random.randint(0,m-1)
        while indice_utilise[indice_aleatoire[i]]:
            indice_aleatoire[i]=random.randint(0,m-1)
        indice_utilise[indice_aleatoire[i]] = True
    for i in range(m):
        population[i] = population_copie[indice_aleatoire[i]]


def algo_gen():
    for gen in tqdm(range(0, nbgen)):
        
        croisement_selectif()
        mutation()
        # Mise à jour des coûts :
        for i in range(m):
            costs[i]=calculate_cost(population[i])
        indice_best_cost = np.argsort(costs)[0]
        if gen == 0:
            bcosts[gen] = costs[indice_best_cost]
            bsolution[gen] = population[indice_best_cost]
        else:
            if costs[indice_best_cost] < bcosts[gen-1]:
                bsolution[gen],bcosts[gen] = recherche_locale_2opt(population[indice_best_cost])
            else:
                bcosts[gen] = bcosts[gen-1]
                bsolution[gen] = bsolution[gen-1]
            population[np.argsort(costs)[-1]]=bsolution[gen]
        avgcosts[gen] = np.average(costs)        
        # Shuffle pour diversifier population
        shuffle_population()
    return
    
def execution_unique():
    initialize()
    
    global bcosts; bcosts=np.zeros(nbgen,dtype=float)
    global bsolution; bsolution = np.zeros([nbgen,n],dtype=int)
    global avgcosts; avgcosts=np.zeros(nbgen,dtype=float)
    
    #__________________ Sauvegarde population initiale __________________
    population_initiale = population[:].copy()
    tab_init =[]
    for i in range(m):
        tab_init.append([population_initiale[i],costs[i]])
    df_init = pd.DataFrame(data=tab_init,columns=['population','coût'])
    #_____________________________________________________________________
    
    #_________________________ Algorithme principal _______________________
    begin = time.time()
    
    algo_gen()

    end=time.time()
    elapsed=end-begin
    #______________________________________________________________________
    
    #__________________ Sauvegarde population finale __________________
    population_finale = population[:].copy()
    tab_fin =[]
    for i in range(m):
        costs[i]=calculate_cost(population_finale[i])
        tab_fin.append([population_finale[i],costs[i]])
    df_fin = pd.DataFrame(data=tab_fin,columns=['population','coût'])
    #_____________________________________________________________________ 
    
    
    
    
    #__________________ Post-processing _____________________________
    
    pd.DataFrame({'best': bcosts,'avg': avgcosts}).plot()
    plt.savefig('BestAndAvg.png')
    pd.DataFrame(avgcosts-bcosts).plot()
    pd.DataFrame(bcosts).plot()
    print("\nMeilleure solution : ",bcosts[nbgen-1])
    print("Meilleure route : ", bsolution[nbgen-1])   
    print('Temps d\'exécution : {:.3f}s - temps moy \'exécution d\'une itération : {:.3f} '.format(elapsed,elapsed/nbgen))
    
    
    writer = pd.ExcelWriter('Résultats_GA.xlsx', engine='xlsxwriter')
    
    workbook  = writer.book
    worksheet = workbook.add_worksheet('Résultat')
    worksheet.set_column(0, 1, 50)
    worksheet.write('A1',"Résultat de l'exécution de l'algorithme génétique")
    worksheet.write('A2',"Heure")
    worksheet.write('B2',time.strftime("%H:%M:%S"))
    worksheet.write('A4',"Meilleur coût")
    worksheet.write('B4',bcosts[nbgen-1])
    worksheet.write('A5',"Meilleure route")
    worksheet.write('B5',str(bsolution[nbgen-1]))
    worksheet.write('A7',"Temps d'exécution")
    worksheet.write('B7',elapsed)
    worksheet.write('A8',"Temps moy d'exécution d'une génération")
    worksheet.write('B8',elapsed/nbgen)
    
    # Écrire dans deux feuilles
    df_init.to_excel(writer, sheet_name='Pop_Initiale')
    df_fin.to_excel(writer, sheet_name='Pop_Finale')

    # Insertion d'une feuille avec le graphe
    worksheet = workbook.add_worksheet('Graphe')
    worksheet.insert_image('B2', 'BestAndAvg.png')
    
    writer.close()
    
    print('\n\n*** Les résultats détaillés se trouvent dans le document Excel Résultats_GA.xlsx ***')
    
    #_____________________________________________________________________ 
    
    return

def execution_multiple():
    '''
    Procédure pour lancer des groupes d'exécutions
    (Utilisée pour le paramétrage)

    '''
    tailles_instances = [20,40,100]
    # 3 instances pour chaque taille
    resultats_total = []
    for dimension in range(3):
        for instance in range(3):
            fname = 'Inst'+str(tailles_instances[dimension])+'_'+str((instance+1))+'.xlsx'
            lecteur(fname)
            print("\nDimension ",tailles_instances[dimension],", instance",instance+1)
            for execution in range(5):
                initialize()
                
                global bcosts; bcosts=np.zeros(nbgen,dtype=float)
                global bsolution; bsolution = np.zeros([nbgen,n],dtype=int)
                global avgcosts; avgcosts=np.zeros(nbgen,dtype=float)
                
                begin = time.time()
                
                algo_gen()

                end=time.time()
                elapsed=end-begin
                
                resultats = [tailles_instances[dimension],instance+1,fname,execution+1,bcosts[nbgen-1],str(bsolution[nbgen-1]),elapsed,elapsed/nbgen]
                resultats_total.append(resultats)

    df_final = pd.DataFrame(resultats_total,columns=['Dimension','Numéro instance','Nom du fichier','Exécution','Meilleure solution','Meilleure route','Temps de résolution','Temps moyen par génération'])
    df_final.to_excel("Resultats_tests_GA.xlsx")
    print('\nLes résultats se trouvent dans le fichier Resultats_tests_GA.xlsx')
    return
                
                
    
    
if __name__ == '__main__':
    main()




